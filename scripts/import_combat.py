#!/usr/bin/env python3
"""
Combat competition history import — Placer Robotics Hub.

Reads the combat history spreadsheet (one row per bot per event) and upserts
combat_event / combat_bot / combat_result into Supabase, all as source='manual'.
This is how historical combat results (SBB, early events) get captured before
any platform integration (Challonge/RCE/Impact) exists. See
docs/combat-results-capture.md.

Accepts .xlsx (first worksheet) or .csv. Columns (header row required, any
order — matched by name):
    event_slug, event_name, date, location, series, bot, weight_class,
    placement, wins, losses, ko_wins, award, notes

    event_slug   stable id, e.g. 'sbb-2025-11' (reused across a bot's rows for
                 the same event so one combat_event row is created per event)
    series       SBB | IRL | NHRL | other
    weight_class plastic_ant | antweight | 15lb | beetleweight
    bot          bot display name (bot_slug is derived by slugifying this)

Usage:
    python scripts/import_combat.py --template history.xlsx   # write a fillable template
    python scripts/import_combat.py history.xlsx              # dry run (default)
    python scripts/import_combat.py history.xlsx --live       # perform upserts

Environment variables (loaded from .env or the process environment):
    SUPABASE_URL                Supabase project URL
    SUPABASE_SERVICE_ROLE_KEY   service-role key (bypasses RLS — server side only)
"""
import argparse
import csv
import re
import sys

try:
    from dotenv import load_dotenv
except ModuleNotFoundError:  # optional — only needed when reading env from a .env file
    def load_dotenv():
        return False

COLUMNS = [
    "event_slug", "event_name", "date", "location", "series", "bot",
    "weight_class", "placement", "wins", "losses", "ko_wins", "award", "notes",
]
SERIES_VALUES = ["SBB", "IRL", "NHRL", "other"]
WEIGHT_VALUES = ["plastic_ant", "antweight", "15lb", "beetleweight"]


def get_supabase():
    from supabase import create_client
    import os

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required")
    return create_client(url, key)


def slugify(name):
    s = re.sub(r"[^a-z0-9]+", "-", name.strip().lower()).strip("-")
    return s or "bot"


def to_int(v):
    v = (v or "").strip()
    return int(float(v)) if v else None  # float() first: Excel numerics arrive as "1.0"


class Stats:
    def __init__(self):
        self.rows_processed = 0
        self.events_upserted = 0
        self.bots_upserted = 0
        self.results_upserted = 0
        self.errors = []   # list[(row_number, message)]

    def error(self, row_no, message):
        self.errors.append((row_no, message))

    def report(self):
        print("\n" + "=" * 60)
        print("Combat history import report")
        print("=" * 60)
        print(f"  Rows processed    : {self.rows_processed}")
        print(f"  Events upserted   : {self.events_upserted}")
        print(f"  Bots upserted     : {self.bots_upserted}")
        print(f"  Results upserted  : {self.results_upserted}")
        print(f"  Errors            : {len(self.errors)}")
        if self.errors:
            print("\n  --- Errors (rows skipped) ---")
            for row_no, message in self.errors:
                print(f"      row {row_no}: {message}")
        print("=" * 60 + "\n")
        return len(self.errors)


REQUIRED_COLS = ["event_slug", "series", "bot", "weight_class"]


def read_rows(path):
    """Yield (fieldnames, iterator of dict rows) from a .csv or .xlsx file."""
    if path.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        ws = load_workbook(path, read_only=True, data_only=True).worksheets[0]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise RuntimeError("spreadsheet is empty")
        fieldnames = [str(h).strip() if h is not None else "" for h in header]

        def cell_str(v):
            import datetime

            if v is None:
                return ""
            if isinstance(v, (datetime.datetime, datetime.date)):
                return v.strftime("%Y-%m-%d")
            if isinstance(v, float) and v.is_integer():
                return str(int(v))  # Excel stores whole numbers as floats
            return str(v).strip()

        def gen():
            for values in rows:
                if values is None or all(v is None or str(v).strip() == "" for v in values):
                    continue  # skip fully blank rows
                yield {name: cell_str(v) for name, v in zip(fieldnames, values) if name}

        return fieldnames, gen()

    f = open(path, newline="", encoding="utf-8-sig")  # noqa: SIM115 — generator keeps it open
    reader = csv.DictReader(f)

    def gen_csv():
        with f:
            yield from reader

    return reader.fieldnames or [], gen_csv()


def write_template(path):
    """Write a fillable .xlsx template: headers, dropdowns for the enum
    columns, and one worked example row (delete it before importing)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "combat history"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B2544")
    ws.append([
        "sbb-2025", "Sacramento Bot Battles 2025", "2025-04-12", "Roseville, CA",
        "SBB", "Lunar Eclipse", "15lb", 1, 4, 0, 2,
        "California BotsIQ 15lb Champion", "example row — delete me",
    ])

    dv_series = DataValidation(type="list", formula1=f'"{",".join(SERIES_VALUES)}"', allow_blank=True)
    dv_weight = DataValidation(type="list", formula1=f'"{",".join(WEIGHT_VALUES)}"', allow_blank=True)
    ws.add_data_validation(dv_series)
    ws.add_data_validation(dv_weight)
    series_col = get_column_letter(COLUMNS.index("series") + 1)
    weight_col = get_column_letter(COLUMNS.index("weight_class") + 1)
    dv_series.add(f"{series_col}2:{series_col}500")
    dv_weight.add(f"{weight_col}2:{weight_col}500")

    widths = {"event_slug": 16, "event_name": 34, "date": 12, "location": 18,
              "series": 9, "bot": 20, "weight_class": 13, "award": 28, "notes": 30}
    for name, width in widths.items():
        ws.column_dimensions[get_column_letter(COLUMNS.index(name) + 1)].width = width

    wb.save(path)
    print(f"Template written to {path} — one row per bot per event. Dates as YYYY-MM-DD.")
    print("The example row shows the format; delete it before importing.")


def import_file(path, db, live, limit=None):
    stats = Stats()
    seen_events = set()
    seen_bots = set()

    fieldnames, rows = read_rows(path)
    missing = [c for c in REQUIRED_COLS if c not in fieldnames]
    if missing:
        raise RuntimeError(f"file is missing required column(s): {', '.join(missing)}")

    for row_no, row in enumerate(rows, start=2):  # header is row 1
        if limit is not None and stats.rows_processed >= limit:
            break
        stats.rows_processed += 1

        event_slug = (row.get("event_slug") or "").strip()
        bot_name = (row.get("bot") or "").strip()
        weight_class = (row.get("weight_class") or "").strip()
        series = (row.get("series") or "").strip()
        if not (event_slug and bot_name and weight_class and series):
            stats.error(row_no, "missing event_slug/bot/weight_class/series")
            continue
        bot_slug = slugify(bot_name)

        event_payload = {
            "event_slug": event_slug,
            "name": (row.get("event_name") or event_slug).strip(),
            "event_date": (row.get("date") or "").strip() or None,
            "location": (row.get("location") or "").strip() or None,
            "series": series,
            "source": "manual",
        }
        bot_payload = {
            "bot_slug": bot_slug,
            "name": bot_name,
            "weight_class": weight_class,
        }
        result_payload = {
            "event_slug": event_slug,
            "bot_slug": bot_slug,
            "weight_class": weight_class,
            "placement": to_int(row.get("placement")),
            "wins": to_int(row.get("wins")) or 0,
            "losses": to_int(row.get("losses")) or 0,
            "ko_wins": to_int(row.get("ko_wins")),
            "award": (row.get("award") or "").strip() or None,
            "notes": (row.get("notes") or "").strip() or None,
            "source": "manual",
        }

        if not live:
            continue

        try:
            if event_slug not in seen_events:
                db.table("combat_event").upsert(event_payload, on_conflict="event_slug").execute()
                seen_events.add(event_slug)
                stats.events_upserted += 1
            if bot_slug not in seen_bots:
                db.table("combat_bot").upsert(bot_payload, on_conflict="bot_slug").execute()
                seen_bots.add(bot_slug)
                stats.bots_upserted += 1
            db.table("combat_result").upsert(
                result_payload, on_conflict="event_slug,bot_slug,weight_class"
            ).execute()
            stats.results_upserted += 1
        except Exception as exc:  # noqa: BLE001
            stats.error(row_no, str(exc))

    return stats


def main(argv=None):
    parser = argparse.ArgumentParser(description="Import combat competition history (.xlsx or .csv) into Supabase")
    parser.add_argument("path", help="path to the history spreadsheet (.xlsx or .csv)")
    parser.add_argument("--template", action="store_true", help="write a fillable .xlsx template to PATH and exit")
    parser.add_argument("--live", action="store_true", help="perform writes (default: dry run)")
    parser.add_argument("--limit", type=int, default=None, help="process only the first N data rows")
    args = parser.parse_args(argv)

    if args.template:
        write_template(args.path)
        return 0

    load_dotenv()

    if args.live:
        try:
            db = get_supabase()
        except Exception as exc:  # noqa: BLE001
            print(f"ERROR connecting to Supabase: {exc}", file=sys.stderr)
            return 2
    else:
        db = None

    try:
        stats = import_file(args.path, db, live=args.live, limit=args.limit)
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    if not args.live:
        print(f"\nDRY RUN — parsed {stats.rows_processed} row(s), 0 written. Pass --live to write.")
    error_count = stats.report()
    return 1 if error_count else 0


if __name__ == "__main__":
    sys.exit(main())
