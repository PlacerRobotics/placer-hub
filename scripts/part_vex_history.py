#!/usr/bin/env python3
"""
PART VEX Competition History — sync job for Placer Robotics Hub.

Pulls PART's historical VEX competition record (Worlds qualifications, banner
awards, elimination-round depth, State/Region titles) from the RobotEvents API
(events.vex.com — www.robotevents.com is dead) and either builds the styled
xlsx brag-book report (ad hoc, local) or upserts the vex_* Supabase tables
(scheduled sync). See docs/vex-stats-integration.md.

Captures three categories:
  - V5RC   : PART's V5 teams (295*, 81818*) — floored at the 2018-19 season
             (VEX reused the "295" numbers for other orgs before PART existed)
  - VIQRC  : PART's IQ teams (295*) — floored at 2018-19
  - 9537   : Cyber Cowboys (Willma Cavitt JH, Middle School V5) — a SEPARATE
             program (is_part=false), full history from whenever RobotEvents
             has it (as far back as 2013-14). Never blended into PART totals.

Worlds awards are NOT in the events API (awards_finalized=false for Worlds),
so they're scraped from the public event awards HTML page and merged in.

Usage:
    python scripts/part_vex_history.py                        # ad hoc: build xlsx, full history, local pickle cache
    VEX_REFRESH=1 python scripts/part_vex_history.py           # ad hoc: force a fresh full pull
    python scripts/part_vex_history.py --backfill --to supabase --live
                                                                 # one-time seed: full history -> Supabase
    python scripts/part_vex_history.py --season current --to supabase --live
                                                                 # scheduled (2x/day): current season only -> Supabase

Environment variables (loaded from .env or the process environment):
    VEX_EVENTS_TOKEN            RobotEvents API bearer token (events.vex.com)
    SUPABASE_URL                Supabase project URL          (only for --to supabase)
    SUPABASE_SERVICE_ROLE_KEY   service-role key, bypasses RLS (only for --to supabase)

Requires: pip install -r scripts/requirements.txt
"""
import argparse
import os
import pickle
import re
import string
import sys
import time
from collections import defaultdict
from datetime import date

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "https://events.vex.com/api/v2"
HTML_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml", "Accept-Language": "en-US,en;q=0.9",
}
SLUG = {"V5RC": "vex-robotics-competition", "VIQRC": "vex-iq-competition"}
CACHE = "part_vex_cache.pkl"

# ── TEAM LISTS ─────────────────────────────────────────────────────────────────
RENAME = {"81818X": "295X", "81818A": "295A", "81818S": "295S"}

V5_TEAMS = [
    "81818X", "81818A", "81818S",
    "295X", "295A", "295S", "295W", "295M",
    "295E", "295F", "295B", "295C", "295D",
    "295G", "295H", "295K",
    "295N", "295P", "295R", "295T",
    "295V", "295Y", "295Z",
]
IQ_TEAMS = [
    "295A", "295B", "295C", "295D", "295E", "295F", "295G", "295H", "295J",
    "295K", "295M", "295N", "295P", "295R", "295S", "295T", "295U",
    "295V", "295W", "295X", "295Y", "295Z",
]

# ── KEYWORDS ───────────────────────────────────────────────────────────────────
WORLDS_KW = ["world championship", "vex world", "worlds"]
# A state/regional CHAMPIONSHIP: name has "championship" AND one of these qualifiers.
STATE_QUAL_KW = ["state", "regional", "region 2", "region-2"]
# Practice/scrimmage/prep events carry championship/worlds words but aren't the real thing.
EXCLUDE_KW = ["practice", "scrimmage", "preparation", "showcase", "prep "]
# Prestigious ("banner") award titles (substring, case-insensitive).
PRESTIGE_KW = ["excellence", "tournament champion", "teamwork champion",
               "design award", "robot skills champion"]


# ── AUTH ───────────────────────────────────────────────────────────────────────
def get_vex_token():
    token = os.environ.get("VEX_EVENTS_TOKEN")
    if not token:
        raise RuntimeError("VEX_EVENTS_TOKEN is required (events.vex.com API token)")
    return token


def get_supabase():
    from supabase import create_client

    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required for --to supabase")
    return create_client(url, key)


# ── API HELPERS ────────────────────────────────────────────────────────────────
def get(url, headers, params=None):
    for attempt in range(5):
        try:
            r = requests.get(url, headers=headers, params=params, timeout=20)
            if r.status_code == 429:  # rate limited — honor Retry-After, back off
                wait = int(r.headers.get("Retry-After", 0)) or (5 * (attempt + 1))
                print(f"    [429 rate-limited, waiting {wait}s]")
                time.sleep(wait)
                continue
            r.raise_for_status()
            return r.json()
        except Exception:
            if attempt == 4:
                raise
            time.sleep(3 * (attempt + 1))


def paginate(url, headers, params=None):
    out = []
    while url:
        d = get(url, headers, params)
        out.extend(d.get("data", []))
        params = None
        url = d.get("meta", {}).get("next_page_url")
        time.sleep(0.25)
    return out


def find_all_team_ids(headers, num, program_code=None):
    d = get(f"{BASE}/teams", headers, {"number[]": num, "per_page": 10})
    teams = d.get("data", [])
    if program_code:
        teams = [t for t in teams if (t.get("program") or {}).get("code", "") == program_code]
    return [(t["id"], (t.get("program") or {}).get("code", "?")) for t in teams]


def discover_9537(headers):
    """Find every 9537* registration (bare 9537 was remapped to 9537A). V5 only."""
    nums = []
    for n in ["9537"] + [f"9537{c}" for c in string.ascii_uppercase]:
        d = get(f"{BASE}/teams", headers, {"number[]": n, "per_page": 10})
        for t in d.get("data", []):
            if (t.get("program") or {}).get("code") == "V5RC" and t.get("number"):
                nums.append(t.get("number"))
    return sorted(set(nums))


def safe(v):
    return v if isinstance(v, str) else ""


# ── SEASON WINDOWING (for --season current: date-bound the /events fetch) ──────
def current_season_start_year():
    """Season runs June 1 - May 31 (CLAUDE.md). Returns the season's START year."""
    today = date.today()
    return today.year if today.month >= 6 else today.year - 1


def season_window(season_start_year):
    """ISO start/end covering a season, with a small buffer past May 31 for
    late-June Worlds / timezone slop. Same param names/shape as the site's
    lib/vexevents.ts syncVexEvents (start=/end= on /teams/{id}/events)."""
    start = f"{season_start_year}-06-01T00:00:00.000Z"
    end = f"{season_start_year + 1}-06-15T00:00:00.000Z"
    return start, end


# ── WORLDS MATCHES (elimination depth) ──────────────────────────────────────────
# Round codes seen at worlds: 1=Practice 2=Qualification 6=Ro16 3=QF 4=SF 5=Final.
# Codes are NOT ordered by bracket depth, so map each to an explicit (name, rank).
STAGE = {1: ("Practice", -1), 2: ("Qualification", 0),
         6: ("Round of 16", 1), 3: ("Quarterfinal", 2), 4: ("Semifinal", 3), 5: ("Final", 4)}


def stage_of(round_code):
    return STAGE.get(round_code, ("Elimination round", 1))  # unknown >2 code -> generic elim


_MATCH_CACHE = {}


def worlds_rounds(headers, tid, event_id):
    key = (tid, event_id)
    if key in _MATCH_CACHE:
        return _MATCH_CACHE[key]
    try:
        ms = paginate(f"{BASE}/teams/{tid}/matches", headers, {"event[]": event_id, "per_page": 250})
        rounds = {m.get("round") for m in ms if m.get("round") is not None}
    except Exception as e:
        print(f"    [matches fetch failed tid={tid} ev={event_id}: {e}]")
        rounds = set()
    _MATCH_CACHE[key] = rounds
    return rounds


def deepest(rounds):
    """Return (stage_name, rank) for the deepest round reached."""
    best = ("—", -99)
    for r in rounds:
        nm, rk = stage_of(r)
        if rk > best[1]:
            best = (nm, rk)
    return best


# ── WORLDS AWARDS (HTML scrape — API omits them, awards_finalized=false) ────────
_SCRAPE_CACHE = {}


def scrape_worlds_awards(sku, program):
    """Return {TEAMNUMBER: [award_title, ...]} for a worlds event's public awards page."""
    if sku in _SCRAPE_CACHE:
        return _SCRAPE_CACHE[sku]
    slug = SLUG.get(program, "vex-robotics-competition")
    url = f"https://events.vex.com/robot-competitions/{slug}/{sku}.html"
    out = defaultdict(list)
    try:
        html = requests.get(url, headers=HTML_HEADERS, timeout=25).text
        rows = re.findall(r"<tr>\s*<td>(.*?)</td>\s*<td>(.*?)</td>\s*<td>(.*?)</td>"
                           r"\s*<td>(.*?)</td>\s*<td>(.*?)</td>\s*</tr>", html, re.S)
        for a, n, _nm, _o, _l in rows:
            title = re.sub("<.*?>", "", a).strip()
            num = re.sub("<.*?>", "", n).strip().upper()
            if title and num and re.match(r"^\d+[A-Z]?$", num):
                out[num].append(title)
    except Exception as e:
        print(f"    [scrape fail {sku}: {e}]")
    _SCRAPE_CACHE[sku] = out
    return out


# ── CLASSIFICATION ─────────────────────────────────────────────────────────────
def is_worlds_name(nm):
    n = (nm or "").lower()
    if any(x in n for x in EXCLUDE_KW):
        return False
    return any(k in n for k in WORLDS_KW)


def classify(ev):
    loc = ev.get("location") or {}
    # region deliberately excluded — it is "California" for every NorCal event.
    name = (safe(ev.get("name")) + " " + safe(loc.get("venue"))).lower()
    if any(x in name for x in EXCLUDE_KW):
        return False, False, ""
    is_w = any(k in name for k in WORLDS_KW)
    is_s = (not is_w) and ("championship" in name) and any(k in name for k in STATE_QUAL_KW)
    scope = "Region" if (is_s and ("region" in name)) else ("State" if is_s else "")
    return is_w, is_s, scope


def get_season(ev):
    """Season START year as an int (2018-2019 season -> 2018)."""
    m = re.search(r'20(\d\d)-20(\d\d)', safe((ev.get("season") or {}).get("name")))
    if m:
        return 2000 + int(m.group(1))
    st = safe(ev.get("start"))
    if len(st) >= 7:
        yr, mo = int(st[:4]), int(st[5:7])
        return yr if mo >= 6 else yr - 1
    m2 = re.search(r'RE-\w+-(\d\d)-', safe(ev.get("sku")))
    if m2:
        return 2000 + int(m2.group(1))
    return 0


def season_label(s):
    return f"{s}-{(s + 1) % 100:02d}"      # 2018 -> "2018-19"


def canon(n, rename):
    return rename.get(n, n)


def is_prestige(title):   # "banner" awards — the ones that earn a hanging banner
    t = title.lower()
    return any(k in t for k in PRESTIGE_KW)


def banner_type(title):
    t = title.lower()
    if "excellence" in t:
        return "Excellence"
    if "tournament champion" in t or "teamwork champion" in t:
        return "Tournament Champions"
    if "design award" in t:
        return "Design"
    if "robot skills champion" in t:
        return "Robot Skills"
    return None


def is_banner(title, is_worlds, is_states):
    """A BANNER is earned only at a Championship event (Kevin's definition):
    the banner-list titles (Excellence / Tournament (or IQ Teamwork) Champions /
    Design / Robot Skills Champion) at a Region/State Championship, or —
    at Worlds — event-level Excellence / Tournament Champions / Robot Skills
    only. Design is EXCLUDED at Worlds (it's a division-level award there).
    The same titles at a regular tournament/league are NOT banners."""
    bt = banner_type(title)
    if bt is None:
        return False
    if is_states:
        return True
    if is_worlds:
        return bt != "Design"
    return False


def tier_rank(is_w, is_s, is_p):
    if is_w:
        return 3            # any award at worlds
    if is_s and is_p:
        return 2            # banner award at a state/region championship
    if is_p:
        return 1            # banner-list title at a regular event (NOT a banner)
    return 0


TIER_LABEL = {3: "Worlds", 2: "State/Region", 1: "Major award", 0: ""}


def norm_title(t):
    return re.sub(r"\(.*?\)", "", t).strip().lower()


# ── PULL ───────────────────────────────────────────────────────────────────────
def new_node():
    return {"events": [], "awards": [], "made_states": False, "state_scopes": set(),
            "made_worlds": False, "worlds_elim": False, "worlds_semi": False,
            "worlds_final": False, "worlds_stage": "", "worlds_awards": [], "wskus": set()}


def pull(headers, team_numbers, program_code, label, rename, min_season=0, only_season=None):
    """only_season, if set, restricts BOTH the /events fetch (date-bounded, like
    lib/vexevents.ts) and everything kept downstream to exactly that season —
    this is what makes --season current cheap: old seasons are never re-fetched
    or re-scraped. min_season is the (unrelated) PART-founding floor used only
    when only_season is None (backfill / full history)."""
    scope_note = f", season {season_label(only_season)} only" if only_season is not None \
        else (f", seasons >= {season_label(min_season)}" if min_season else "")
    print(f"\n{'=' * 58}\n{label} (program={program_code}) — {len(team_numbers)} team numbers{scope_note}\n{'=' * 58}")
    data = {}

    ev_params = {"per_page": 250}
    if only_season is not None:
        s0, s1 = season_window(only_season)
        ev_params["start"] = s0
        ev_params["end"] = s1

    def in_scope(s):
        if only_season is not None:
            return s == only_season
        return s >= max(min_season, 2010)

    for raw in sorted(set(team_numbers)):
        cn = canon(raw, rename)
        print(f"  {raw} ({cn})...", end=" ", flush=True)
        team_ids = find_all_team_ids(headers, raw, program_code=program_code)
        if not team_ids:
            print("NOT FOUND")
            continue

        all_events = []
        all_awards = []
        wrounds_by_season = defaultdict(set)   # season -> set(round codes) at worlds
        wsku_by_season = defaultdict(set)      # season -> set((sku,program))
        for tid, prog in team_ids:
            evs = paginate(f"{BASE}/teams/{tid}/events", headers, dict(ev_params))
            aws = paginate(f"{BASE}/teams/{tid}/awards", headers, {"per_page": 250})
            all_events.extend(evs)
            all_awards.extend(aws)
            for ev in evs:
                if is_worlds_name(ev.get("name")):
                    s = get_season(ev)
                    if not in_scope(s):
                        continue
                    sku = safe(ev.get("sku"))
                    wrounds_by_season[s] |= worlds_rounds(headers, tid, ev.get("id"))
                    if sku:
                        wsku_by_season[s].add((sku, program_code))

        # dedup events by sku, awards by (title,event id)
        seen = set()
        events = []
        for ev in all_events:
            k = ev.get("sku") or ev.get("id")
            if k not in seen:
                seen.add(k)
                events.append(ev)
        seen = set()
        awards = []
        for aw in all_awards:
            k = (aw.get("title", ""), (aw.get("event") or {}).get("id", ""))
            if k not in seen:
                seen.add(k)
                awards.append(aw)
        print(f"ids={[t for t, _ in team_ids]} -> {len(events)} events, {len(awards)} awards")

        if cn not in data:
            data[cn] = defaultdict(new_node)

        # events -> build id->season/flags map for award lookup
        emap = {}
        for ev in events:
            s = get_season(ev)
            if not in_scope(s):
                continue
            is_w, is_s, scope = classify(ev)
            loc = ev.get("location") or {}
            eid = ev.get("id")
            if eid is not None:
                emap[eid] = {"season": s, "is_w": is_w, "is_s": is_s, "scope": scope}
            nd = data[cn][s]
            nd["events"].append({"name": safe(ev.get("name")), "sku": safe(ev.get("sku")),
                "loc": f"{safe(loc.get('city'))} {safe(loc.get('region'))}".strip(),
                "start": safe(ev.get("start", ""))[:10], "is_worlds": is_w, "is_states": is_s})
            if is_w:
                nd["made_worlds"] = True
            if is_s:
                nd["made_states"] = True
                nd["state_scopes"].add(scope)

        # worlds elimination depth per season (from matches)
        for s, rounds in wrounds_by_season.items():
            nd = data[cn][s]
            nd["wskus"] |= wsku_by_season.get(s, set())
            stage, rank = deepest(rounds)
            if rank >= 0:
                nd["worlds_stage"] = stage
            elif nd["made_worlds"]:
                nd["worlds_stage"] = "Qualified (no match data)"
            nd["worlds_elim"] = nd["worlds_elim"] or rank >= 1
            nd["worlds_semi"] = nd["worlds_semi"] or rank >= 3
            nd["worlds_final"] = nd["worlds_final"] or rank >= 4

        # API awards -> inherit season/flags from the event map
        for aw in awards:
            ev_obj = aw.get("event") or {}
            ev_name = safe(ev_obj.get("name"))
            aw_name = safe(aw.get("title"))
            if not aw_name:
                continue
            meta = emap.get(ev_obj.get("id"))
            if meta:
                s = meta["season"]
                is_w = meta["is_w"]
                is_s = meta["is_s"]
                scope = meta["scope"]
            else:
                m = re.search(r'RE-\w+-(\d\d)-', safe(ev_obj.get("code")))
                s = 2000 + int(m.group(1)) if m else 0
                is_w = is_worlds_name(ev_name)
                is_s = False
                scope = ""
            if not in_scope(s):
                continue
            is_p = is_prestige(aw_name)
            data[cn][s]["awards"].append({"award": aw_name, "event": ev_name,
                "event_sku": safe(ev_obj.get("code")) or None,
                "is_worlds": is_w, "is_states": is_s, "state_scope": scope, "is_prestige": is_p,
                "is_banner": is_banner(aw_name, is_w, is_s),
                "tier": tier_rank(is_w, is_s, is_p), "source": "api"})
            if is_w:
                data[cn][s]["worlds_awards"].append(f"{aw_name} @ {ev_name}")

    # ── worlds-award HTML scrape (merge awards the API omits) ────────────────────
    inv = defaultdict(set)   # cn -> raw numbers that canon to it (for scrape matching)
    for raw in team_numbers:
        inv[canon(raw, rename)].add(raw.upper())
    for cn in data:
        inv[cn].add(cn.upper())
    scraped_added = 0
    for cn in data:
        for s in list(data[cn].keys()):
            nd = data[cn][s]
            for sku, prog in nd["wskus"]:
                page = scrape_worlds_awards(sku, prog)
                for num in inv[cn]:
                    for title in page.get(num, []):
                        nt = norm_title(title)
                        if any(a["is_worlds"] and norm_title(a["award"]) == nt for a in nd["awards"]):
                            continue
                        is_p = is_prestige(title)
                        nd["awards"].append({"award": title, "event": f"VEX Worlds {season_label(s)}",
                            "event_sku": sku,
                            "is_worlds": True, "is_states": False, "state_scope": "", "is_prestige": is_p,
                            "is_banner": is_banner(title, True, False),
                            "tier": 3, "source": "worlds-html"})
                        nd["worlds_awards"].append(f"{title} @ VEX Worlds")
                        scraped_added += 1
    if scraped_added:
        print(f"  + {scraped_added} worlds award(s) merged from HTML")
    return data


# ── AGGREGATION ─────────────────────────────────────────────────────────────────
def seasons_of(data):
    return sorted({s for t in data for s in data[t] if data[t][s]["events"]})


def agg(data, seasons):
    A = dict(teams=0, seasons=0, events=0, awards=0, prestige=0, banners=0, banner_state=0,
             state_aw=0, region_aw=0, state_appear=0, wq=0, wq_seasons=0,
             welim=0, wsemi=0, wfinal=0, waws=0)
    active = set()
    for t in data:
        tactive = False
        for s in seasons:
            nd = data[t].get(s)
            if not nd or not nd["events"]:
                continue
            tactive = True
            active.add(s)
            A["events"] += len(nd["events"])
            A["awards"] += len(nd["awards"])
            A["prestige"] += sum(1 for a in nd["awards"] if a["is_prestige"])
            A["banners"] += sum(1 for a in nd["awards"] if a["is_banner"])
            A["banner_state"] += sum(1 for a in nd["awards"] if a["is_banner"] and a["is_states"])
            A["state_aw"] += sum(1 for a in nd["awards"] if a["is_states"] and a["state_scope"] == "State")
            A["region_aw"] += sum(1 for a in nd["awards"] if a["is_states"] and a["state_scope"] == "Region")
            if nd["made_states"]:
                A["state_appear"] += 1
            if nd["made_worlds"]:
                A["wq"] += 1
            if nd["worlds_elim"]:
                A["welim"] += 1
            if nd["worlds_semi"]:
                A["wsemi"] += 1
            if nd["worlds_final"]:
                A["wfinal"] += 1
            A["waws"] += len(nd["worlds_awards"])
        if tactive:
            A["teams"] += 1
    A["seasons"] = len(active)
    A["wq_seasons"] = len({s for t in data for s in seasons
                            if data[t].get(s) and data[t][s]["made_worlds"]})
    A["wq_teams"] = len({t for t in data for s in seasons
                          if data[t].get(s) and data[t][s]["made_worlds"]})
    return A


# ── EXCEL STYLING ────────────────────────────────────────────────────────────────
GOLD = "E9BE45"
NAVY = "07111F"
LGRAY = "F4F9FF"
GREEN = "D6F4D9"
GOLDL = "FBEFC6"
BLUEL = "E4EDFB"


def h(ws, r, c, v, bg=NAVY, fg="FFFFFF", sz=10):
    x = ws.cell(r, c, v)
    x.font = Font(bold=True, color=fg, size=sz)
    x.fill = PatternFill("solid", fgColor=bg)
    x.alignment = Alignment(horizontal="center", wrap_text=True)


def cel(ws, r, c, v, bg=None, bold=False, align="left", italic=False, fg=None):
    x = ws.cell(r, c, v)
    x.font = Font(bold=bold, italic=italic, color=fg) if fg else Font(bold=bold, italic=italic)
    if bg:
        x.fill = PatternFill("solid", fgColor=bg)
    x.alignment = Alignment(horizontal=align, vertical="center")


# ── SHEETS: per-category ─────────────────────────────────────────────────────────
def awards_sheet(wb, data, label, seasons):
    ws = wb.create_sheet(f"{label} — All Awards")
    ws.freeze_panes = "A2"
    cols = ["Team", "Season", "Award", "Event", "Level", "Banner?", "Worlds?", "State/Region?"]
    for ci, c in enumerate(cols, 1):
        h(ws, 1, ci, c)
    row = 2
    for tn in sorted(data):
        for s in seasons:
            if s not in data[tn]:
                continue
            for aw in sorted(data[tn][s]["awards"], key=lambda a: (-a["tier"], a["award"])):
                lvl = TIER_LABEL[aw["tier"]] if aw["is_worlds"] else (aw["state_scope"] or "")
                bg = GREEN if aw["is_worlds"] else (GOLDL if aw["tier"] >= 1 else (LGRAY if row % 2 == 0 else "FFFFFF"))
                cells = [tn, season_label(s), aw["award"], aw["event"], lvl,
                         "✓" if aw["is_banner"] else "",
                         "✓" if aw["is_worlds"] else "",
                         aw["state_scope"]]
                for ci, v in enumerate(cells, 1):
                    cel(ws, row, ci, v, bg=bg, align="center" if ci >= 5 else "left",
                        bold=(ci == 3 and aw["tier"] >= 2))
                row += 1
    for col, w in {"A": 9, "B": 9, "C": 40, "D": 48, "E": 13, "F": 10, "G": 9, "H": 13}.items():
        ws.column_dimensions[col].width = w


def team_summary(wb, data, label, seasons):
    ws = wb.create_sheet(f"{label} — Team Summary")
    ws.freeze_panes = "B2"
    hdr = ["Team"] + [season_label(s) for s in seasons] + [
        "Tot Events", "Tot Awards", "Banner", "State/Reg Aws",
        "States Quals", "Worlds Quals", "Worlds Elim", "Worlds SF+", "Worlds Aws"]
    for ci, c in enumerate(hdr, 1):
        h(ws, 1, ci, c, sz=9)
    ws.row_dimensions[1].height = 30
    row = 2
    for tn in sorted(data):
        te = ta = tp = tsr = sq = wq = we = wsf = wa = 0
        per = []
        for s in seasons:
            nd = data[tn].get(s)
            if nd and nd["events"]:
                ec = len(nd["events"])
                ac = len(nd["awards"])
                mark = "W" if nd["made_worlds"] else ("S" if nd["made_states"] else "")
                per.append(f"{ec}/{ac}" + (f" {mark}" if mark else ""))
                te += ec
                ta += ac
                tp += sum(1 for a in nd["awards"] if a["is_banner"])
                tsr += sum(1 for a in nd["awards"] if a["is_states"])
                if nd["made_states"]:
                    sq += 1
                if nd["made_worlds"]:
                    wq += 1
                if nd["worlds_elim"]:
                    we += 1
                if nd["worlds_semi"]:
                    wsf += 1
                wa += len(nd["worlds_awards"])
            else:
                per.append("")
        vals = [tn] + per + [te, ta, tp, tsr, sq, wq, we, wsf, wa]
        bg = LGRAY if row % 2 == 0 else "FFFFFF"
        for ci, v in enumerate(vals, 1):
            cel(ws, row, ci, v, bg=bg, align="center" if ci > 1 else "left",
                bold=(isinstance(v, int) and v > 0 and ci > len(seasons) + 1))
        row += 1
    ws.column_dimensions["A"].width = 9
    for i in range(2, len(seasons) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 9
    for i in range(len(seasons) + 2, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10
    note = ws.cell(row + 1, 1, 'Cell = events/awards that season;  "W"=went to Worlds,  "S"=made State/Region championship.')
    note.font = Font(italic=True, color="888888", size=9)
    return ws


def year_summary(wb, data, label, seasons):
    ws = wb.create_sheet(f"{label} — Season Summary")
    cols = ["Season", "# Teams", "Events", "Awards", "Banner", "State/Reg Aws",
            "States Quals", "Worlds Quals", "Worlds Elim", "Worlds SF+", "Worlds Aws", "Worlds Detail"]
    for ci, c in enumerate(cols, 1):
        h(ws, 1, ci, c)
    ws.row_dimensions[1].height = 28
    row = 2
    for s in seasons:
        active = [t for t in data if s in data[t] and data[t][s]["events"]]
        if not active:
            continue
        ev = sum(len(data[t][s]["events"]) for t in active)
        aw = sum(len(data[t][s]["awards"]) for t in active)
        pr = sum(sum(1 for a in data[t][s]["awards"] if a["is_banner"]) for t in active)
        sr = sum(sum(1 for a in data[t][s]["awards"] if a["is_states"]) for t in active)
        sq = sum(1 for t in active if data[t][s]["made_states"])
        wq = sum(1 for t in active if data[t][s]["made_worlds"])
        we = sum(1 for t in active if data[t][s]["worlds_elim"])
        wsf = sum(1 for t in active if data[t][s]["worlds_semi"])
        wa = sum(len(data[t][s]["worlds_awards"]) for t in active)
        det = "; ".join(f"{t} ({data[t][s]['worlds_stage'] or 'Worlds'})"
                         for t in active if data[t][s]["made_worlds"])
        bg = GREEN if wq > 0 else (LGRAY if row % 2 == 0 else "FFFFFF")
        vals = [season_label(s), len(active), ev, aw, pr, sr, sq, wq, we, wsf, wa, det]
        for ci, v in enumerate(vals, 1):
            cel(ws, row, ci, v, bg=bg, align="center" if ci < 12 else "left", bold=(ci == 1))
        row += 1
    for i in range(1, 12):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions["L"].width = 55


# ── SHEETS: cross-category ───────────────────────────────────────────────────────
def headline_sheet(wb, cats):
    ws = wb.create_sheet("⭐ Headline Stats", 0)
    ws.sheet_view.showGridLines = False

    def sec(r, title):
        c = ws.cell(r, 1, title)
        c.font = Font(bold=True, size=13, color=NAVY)
        c.fill = PatternFill("solid", fgColor=GOLD)
        ws.merge_cells(f"A{r}:D{r}")
        return r + 1

    def stat(r, lbl, val, note=""):
        ws.cell(r, 1, lbl).font = Font(size=11)
        c = ws.cell(r, 2, val)
        c.font = Font(bold=True, size=13, color=NAVY)
        c.alignment = Alignment(horizontal="center")
        if note:
            ws.cell(r, 3, note).font = Font(italic=True, color="888888", size=10)
        return r + 1

    def detail(r, line, bold=False, color="000000"):
        ws.cell(r, 1, line).font = Font(size=10, bold=bold, color=color)
        return r + 1

    r = 1
    ws.cell(r, 1, "PART Competitive History — All Time").font = Font(bold=True, size=16, color=NAVY)
    ws.merge_cells(f"A{r}:E{r}")
    r += 2
    for label, data, seasons in cats:
        A = agg(data, seasons)
        r = sec(r, f"{label}   ({season_label(seasons[0])} – {season_label(seasons[-1])})" if seasons else label)
        r = stat(r, "Seasons active", A["seasons"])
        r = stat(r, "Total events", A["events"])
        r = stat(r, "Total awards", A["awards"])
        bcount = defaultdict(int)
        for t in data:
            for s in seasons:
                if s not in data[t]:
                    continue
                for a in data[t][s]["awards"]:
                    if a["is_banner"]:
                        bcount[banner_type(a["award"])] += 1
        r = stat(r, "Banner awards", A["banners"],
                 "championship-level only: State/Region champs, or Worlds event-level Excellence/Tourney Champs/Robot Skills")
        r = detail(r, f'      Excellence {bcount["Excellence"]}  ·  Tournament Champions {bcount["Tournament Champions"]}'
                      f'  ·  Design {bcount["Design"]}  ·  Robot Skills {bcount["Robot Skills"]}', color="555555")
        r = stat(r, "   — of which at a State/Region championship", A["banner_state"])
        r = stat(r, "Banner-list titles at any event", A["prestige"],
                 "same titles at regular tournaments/leagues — NOT banners")
        r = stat(r, "State/Region championship awards", A["state_aw"] + A["region_aw"],
                 f'({A["state_aw"]} all-CA State · {A["region_aw"]} Region)')
        r = stat(r, "State/Region championship appearances", A["state_appear"], "(team-seasons)")
        r = stat(r, "Worlds qualifications (by team)", A["wq"], "each team, each season it qualified")
        r = stat(r, "Distinct teams that reached Worlds", A["wq_teams"])
        r = stat(r, "Seasons with a Worlds qualifier", A["wq_seasons"])
        r = stat(r, "Worlds elimination appearances", A["welim"])
        r = stat(r, "Worlds semifinal-or-better", A["wsemi"])
        r = stat(r, "Worlds finalist appearances", A["wfinal"])
        r = stat(r, "Worlds awards", A["waws"])
        sr = [(s, t, a) for t in data for s in seasons if s in data[t]
              for a in data[t][s]["awards"] if a["is_states"]]
        if sr:
            r += 1
            r = detail(r, "State / Region championship award detail:", bold=True)
            for s, t, a in sorted(sr, key=lambda x: (x[0], x[1], x[2]["award"])):
                flag = "   🚩 BANNER" if a["is_prestige"] else ""
                r = detail(r, f'   {t} ({season_label(s)}): {a["award"]} — {a["state_scope"]}{flag}',
                           bold=bool(a["is_prestige"]), color=(NAVY if a["is_prestige"] else "444444"))
        det = [(t, s, a) for t in data for s in seasons if s in data[t]
               for a in data[t][s]["worlds_awards"]]
        if det:
            r += 1
            r = detail(r, "Worlds award detail:", bold=True)
            for t, s, a in det:
                r = detail(r, f"   {t} ({season_label(s)}): {a}")
        r += 1
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 52


def totals_sheet(wb, cats):
    ws = wb.create_sheet("🏆 Totals", 1)
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "Combined Totals — All Programs").font = Font(bold=True, size=15, color=NAVY)
    ws.merge_cells("A1:F1")
    rows = [("Seasons active", "seasons"), ("Total events", "events"), ("Total awards", "awards"),
            ("Banner awards (championship-level only)", "banners"),
            ("  — of which at a State/Region championship", "banner_state"),
            ("Banner-list titles at any event (not banners)", "prestige"),
            ("State championship awards (all-CA)", "state_aw"),
            ("Region championship awards", "region_aw"),
            ("State/Region championship appearances", "state_appear"),
            ("Worlds qualifications (by team, each season)", "wq"),
            ("Distinct teams that reached Worlds", "wq_teams"),
            ("Seasons with a Worlds qualifier", "wq_seasons"),
            ("Worlds elimination appearances", "welim"), ("Worlds semifinal-or-better", "wsemi"),
            ("Worlds finalist appearances", "wfinal"), ("Worlds awards", "waws"),
            ("Distinct teams", "teams")]
    aggs = [(lbl, agg(data, seasons), seasons, data) for lbl, data, seasons in cats]
    h(ws, 3, 1, "Metric")
    for ci, (lbl, _, _, _) in enumerate(aggs, 2):
        h(ws, 3, ci, lbl)
    h(ws, 3, len(aggs) + 2, "ALL", bg=GOLD, fg=NAVY)
    all_seasons = set()
    all_wq_seasons = set()
    for lbl, A, seasons, data in aggs:
        all_seasons |= {s for t in data for s in seasons if data[t].get(s) and data[t][s]["events"]}
        all_wq_seasons |= {s for t in data for s in seasons if data[t].get(s) and data[t][s]["made_worlds"]}
    for ri, (name, key) in enumerate(rows, 4):
        cel(ws, ri, 1, name, bold=True, bg=(LGRAY if ri % 2 == 0 else "FFFFFF"))
        total = 0
        for ci, (lbl, A, seasons, data) in enumerate(aggs, 2):
            cel(ws, ri, ci, A[key], align="center", bg=(LGRAY if ri % 2 == 0 else "FFFFFF"))
            total += A[key]
        if key == "seasons":
            total = len(all_seasons)
        elif key == "wq_seasons":
            total = len(all_wq_seasons)
        cel(ws, ri, len(aggs) + 2, total, align="center", bold=True, bg=GOLDL)
    ws.column_dimensions["A"].width = 40
    for i in range(2, len(aggs) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 13


def elevated_sheet(wb, cats):
    ws = wb.create_sheet("★ Elevated Awards", 2)
    ws.freeze_panes = "A2"
    cols = ["Tier", "Category", "Team", "Season", "Award", "Event", "Level"]
    for ci, c in enumerate(cols, 1):
        h(ws, 1, ci, c)
    items = []
    for label, data, seasons in cats:
        for t in data:
            for s in seasons:
                if s not in data[t]:
                    continue
                for a in data[t][s]["awards"]:
                    if a["tier"] >= 1:
                        lvl = TIER_LABEL[a["tier"]] if a["is_worlds"] else (a["state_scope"] or "Event")
                        items.append((a["tier"], label, t, s, a["award"], a["event"], lvl))
    items.sort(key=lambda x: (-x[0], x[3], x[1], x[2]))
    row = 2
    for tier, label, t, s, aw, ev, lvl in items:
        stars = "★" * tier
        bg = GREEN if tier == 3 else (GOLDL if tier == 2 else (BLUEL if row % 2 == 0 else "FFFFFF"))
        vals = [stars, label, t, season_label(s), aw, ev, lvl]
        for ci, v in enumerate(vals, 1):
            cel(ws, row, ci, v, bg=bg, align="center" if ci in (1, 4, 7) else "left", bold=(ci == 5 and tier >= 2))
        row += 1
    for col, w in {"A": 7, "B": 8, "C": 9, "D": 9, "E": 40, "F": 46, "G": 13}.items():
        ws.column_dimensions[col].width = w
    note = ws.cell(row + 1, 1, "★★★ = award at Worlds · ★★ = BANNER (banner-list title at a State/Region championship) · ★ = banner-list title at a regular event (not a banner). Banner-list = Excellence · Tournament Champions · Design · Robot Skills; at Worlds only event-level Excellence/Tourney Champs/Robot Skills earn a banner (Design there is division-level).")
    note.font = Font(italic=True, color="888888", size=9)


def worlds_runs_sheet(wb, cats):
    ws = wb.create_sheet("🌎 Worlds Runs", 3)
    ws.freeze_panes = "A2"
    cols = ["Category", "Team", "Season", "Deepest Stage", "Elim?", "SF+?", "Finalist?", "Worlds Awards"]
    for ci, c in enumerate(cols, 1):
        h(ws, 1, ci, c)
    runs = []
    for label, data, seasons in cats:
        for t in data:
            for s in seasons:
                nd = data[t].get(s)
                if nd and nd["made_worlds"]:
                    runs.append((s, label, t, nd))
    runs.sort(key=lambda x: (x[0], x[1], x[2]))
    row = 2
    for s, label, t, nd in runs:
        bg = GOLDL if nd["worlds_semi"] else (GREEN if nd["worlds_elim"] else (LGRAY if row % 2 == 0 else "FFFFFF"))
        vals = [label, t, season_label(s), nd["worlds_stage"] or "Qualified",
                "✓" if nd["worlds_elim"] else "", "✓" if nd["worlds_semi"] else "",
                "✓" if nd["worlds_final"] else "", "; ".join(nd["worlds_awards"])]
        for ci, v in enumerate(vals, 1):
            cel(ws, row, ci, v, bg=bg, align="center" if ci in (3, 5, 6, 7) else "left",
                bold=(ci == 4 and nd["worlds_semi"]))
        row += 1
    for col, w in {"A": 8, "B": 9, "C": 9, "D": 20, "E": 7, "F": 7, "G": 9, "H": 50}.items():
        ws.column_dimensions[col].width = w


def build_workbook(cats, out_path="part_vex_history.xlsx"):
    print("\nBuilding Excel workbook...")
    wb = Workbook()
    wb.remove(wb.active)
    headline_sheet(wb, cats)
    totals_sheet(wb, cats)
    elevated_sheet(wb, cats)
    worlds_runs_sheet(wb, cats)
    for lbl, data, seasons in cats:
        awards_sheet(wb, data, lbl, seasons)
        team_summary(wb, data, lbl, seasons)
        year_summary(wb, data, lbl, seasons)
    wb.save(out_path)
    print(f"\n✓ Saved: {out_path}")
    for s in wb.sheetnames:
        print(f"  • {s}")


# ── SUPABASE SYNC ────────────────────────────────────────────────────────────────
CATEGORY_META = {
    "V5RC": {"category": "v5rc", "program": "vex_v5", "is_part": True},
    "VIQRC": {"category": "viqrc", "program": "vex_iq", "is_part": True},
    "9537": {"category": "cyber9537", "program": "vex_v5", "is_part": False},
}


def to_supabase_rows(cats, only_season):
    """Flatten pull() output into vex_team / vex_award / vex_worlds_run row dicts.
    Teams key on (team_number, program): PART reuses the same numbers across V5
    and IQ ("295A" is two different teams), so number alone collapses categories
    and mis-attributes every award."""
    teams = {}
    awards = []
    runs = []
    for label, data, seasons in cats:
        meta = CATEGORY_META[label]
        prog = meta["program"]
        for tn in data:
            all_seasons = sorted(s for s in data[tn] if data[tn][s]["events"])
            if not all_seasons:
                continue
            team_row = {
                "team_number": tn, "program": prog, "category": meta["category"],
                "is_part": meta["is_part"],
            }
            if only_season is None:
                # Only a backfill run knows the FULL history — a current-season
                # run must NOT overwrite first/last season with just this one.
                team_row["first_season"] = season_label(all_seasons[0])
                team_row["last_season"] = season_label(all_seasons[-1])
            teams[(tn, prog)] = team_row
            for s in all_seasons:
                if only_season is not None and s != only_season:
                    continue
                nd = data[tn][s]
                for a in nd["awards"]:
                    if a["source"] not in ("api", "worlds-html"):
                        continue
                    awards.append({
                        "team_number": tn, "program": prog, "season": season_label(s),
                        "title": a["award"],
                        "event_name": a["event"] or None, "event_sku": a.get("event_sku"),
                        "is_worlds": a["is_worlds"], "scope": (a["state_scope"] or None),
                        "is_banner": a["is_banner"], "banner_type": banner_type(a["award"]),
                        "source": a["source"],
                    })
                if nd["made_worlds"]:
                    sku = sorted(nd["wskus"])[0][0] if nd["wskus"] else None
                    runs.append({
                        "team_number": tn, "program": prog, "season": season_label(s),
                        "event_sku": sku,
                        "deepest_stage": nd["worlds_stage"] or "Qualified",
                        "made_elim": nd["worlds_elim"], "made_semi": nd["worlds_semi"],
                        "made_final": nd["worlds_final"], "source": "api",
                    })
    return list(teams.values()), awards, runs


def push_to_supabase(cats, only_season, live):
    teams, awards, runs = to_supabase_rows(cats, only_season)
    if only_season is not None:
        season_labels = [season_label(only_season)]
        scope_desc = f"current season only ({season_labels[0]})"
    else:
        all_seasons = sorted({s for _, _, seasons in cats for s in seasons})
        season_labels = [season_label(s) for s in all_seasons]
        scope_desc = f"full backfill ({len(season_labels)} seasons)"

    print(f"\n{'=' * 60}\nSupabase sync — {'LIVE' if live else 'DRY RUN'}\n{'=' * 60}")
    print(f"  Scope           : {scope_desc}")
    print(f"  Teams to upsert : {len(teams)}")
    print(f"  Awards to write : {len(awards)}  (source=api/worlds-html)")
    print(f"  Worlds runs     : {len(runs)}")

    if not live:
        print("\n  DRY RUN — pass --live to write. No changes made.")
        return 0

    try:
        db = get_supabase()
    except Exception as exc:
        print(f"ERROR connecting to Supabase: {exc}", file=sys.stderr)
        return 2

    for t in teams:
        db.table("vex_team").upsert(t, on_conflict="team_number,program").execute()

    # Clean re-sync of this run's season scope. Only rows this script owns
    # (source in api/worlds-html) are ever touched — source='manual' rows
    # (hand-entered, e.g. Worlds results the scrape missed) are never deleted.
    for lbl in season_labels:
        db.table("vex_award").delete().eq("season", lbl).in_("source", ["api", "worlds-html"]).execute()
        db.table("vex_worlds_run").delete().eq("season", lbl).in_("source", ["api", "worlds-html"]).execute()
    if awards:
        db.table("vex_award").insert(awards).execute()
    if runs:
        db.table("vex_worlds_run").insert(runs).execute()

    print("\n  ✓ Supabase updated.")
    return 0


# ── MAIN ───────────────────────────────────────────────────────────────────────
def pull_all(headers, only_season=None):
    print("Discovering 9537 (Cyber Cowboys) registrations...")
    nine = discover_9537(headers)
    print(f"  found: {', '.join(nine)}")
    # PART founded 2018 — floor 295 teams at the 2018-19 season (VEX reused the
    # "295" numbers for other orgs before PART). 9537 (Cyber Cowboys) keeps its
    # full history — a separate program, never blended into PART's totals.
    v5 = pull(headers, V5_TEAMS, "V5RC", "V5RC", RENAME, min_season=2018, only_season=only_season)
    iq = pull(headers, IQ_TEAMS, "VIQRC", "VIQRC", {}, min_season=2018, only_season=only_season)
    cc = pull(headers, nine, "V5RC", "9537", {}, only_season=only_season)
    return v5, iq, cc


def main(argv=None):
    parser = argparse.ArgumentParser(description="PART VEX competition history: pull + report or Supabase sync")
    scope = parser.add_mutually_exclusive_group()
    scope.add_argument("--backfill", action="store_true",
                        help="Force a full re-pull covering ALL seasons (slow; run rarely/manually).")
    scope.add_argument("--season", choices=["current"],
                        help="Pull ONLY the current season (fast, safe for the 2x/day scheduled sync).")
    parser.add_argument("--to", choices=["xlsx", "supabase"], default="xlsx",
                         help="Output target: build the styled xlsx report (default) or push to Supabase.")
    parser.add_argument("--live", action="store_true",
                         help="With --to supabase: actually write. Default is dry-run (compute + report only).")
    args = parser.parse_args(argv)

    load_dotenv()

    try:
        token = get_vex_token()
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

    only_season = current_season_start_year() if args.season == "current" else None

    use_cache = (
        args.to == "xlsx" and not args.backfill and args.season is None
        and os.path.exists(CACHE) and not os.environ.get("VEX_REFRESH")
    )
    if use_cache:
        print(f"Loading cached data from {CACHE} (VEX_REFRESH=1 to re-pull)...")
        v5, iq, cc = pickle.load(open(CACHE, "rb"))
    else:
        v5, iq, cc = pull_all(headers, only_season=only_season)
        if args.to == "xlsx" and only_season is None:
            pickle.dump((v5, iq, cc), open(CACHE, "wb"))
            print(f"Cached raw data -> {CACHE}")

    cats = [("V5RC", v5, seasons_of(v5)), ("VIQRC", iq, seasons_of(iq)), ("9537", cc, seasons_of(cc))]
    for lbl, data, seasons in cats:
        if seasons:
            print(f"{lbl}: seasons {season_label(seasons[0])} – {season_label(seasons[-1])}")

    if args.to == "supabase":
        return push_to_supabase(cats, only_season, live=args.live)

    build_workbook(cats)
    return 0


if __name__ == "__main__":
    sys.exit(main())
